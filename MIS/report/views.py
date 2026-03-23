from django.http import JsonResponse
from django.db.models import Sum, Count, Q
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import timedelta
import json
from finance.models import LoanTransaction, Payment
from beneficiaries.models import District, Municipality, Barangay
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
from docx import Document
from docx.shared import Inches, Pt

def api_districts(request):
    districts = District.objects.values('id', 'name').order_by('name')
    return JsonResponse(list(districts), safe=False)

def api_municipalities(request):
    district_id = request.GET.get('district_id')
    municipalities = []
    if district_id:
        municipalities = Municipality.objects.filter(district_id=district_id).values('id', 'name').order_by('name')
    return JsonResponse(list(municipalities), safe=False)

def api_barangays(request):
    municipality_id = request.GET.get('municipality_id')
    barangays = []
    if municipality_id:
        barangays = Barangay.objects.filter(municipality_id=municipality_id).values('id', 'name').order_by('name')
    return JsonResponse(list(barangays), safe=False)
    return JsonResponse(list(barangays), safe=False)

# 🔥 FIXED report_summary_api (date range was backwards)
@require_http_methods(["POST"])
def report_summary_api(request):
    try:
        data = json.loads(request.body)
        period_type = data.get('period', 'monthly')
        district_id = data.get('district')
        municipality_id = data.get('municipality')
        barangay_id = data.get('barangay')
        
        # 🔥 FIXED: Calculate date range FIRST
        end_date = timezone.now().date()
        if period_type == 'weekly':
            start_date = end_date - timedelta(days=6)
        elif period_type == 'monthly':
            start_date = end_date.replace(day=1)
        elif period_type == '6months':
            start_date = end_date - timedelta(days=180)
        elif period_type == '1year':
            start_date = end_date - timedelta(days=365)
        else:
            start_date = end_date - timedelta(days=30)
        
        # 🔥 Base querysets
        loans = LoanTransaction.objects.filter(loan_date__range=[start_date, end_date])
        payments = Payment.objects.filter(payment_date__range=[start_date, end_date])
        
        # 🔥 Apply filters
        if district_id:
            loans = loans.filter(district_id=district_id)
            payments = payments.filter(beneficiary__district_id=district_id)
        
        if municipality_id:
            loans = loans.filter(municipality_id=municipality_id)
            payments = payments.filter(beneficiary__municipality_id=municipality_id)
        
        if barangay_id:
            loans = loans.filter(barangay_id=barangay_id)
            payments = payments.filter(beneficiary__barangay_id=barangay_id)
        
        # 🔥 Summary
        summary = {
            'total_loans': float(loans.aggregate(Sum('amount'))['amount__sum'] or 0),
            'loan_count': loans.count(),
            'total_payments': float(payments.aggregate(Sum('amount'))['amount__sum'] or 0),
            'payment_count': payments.count(),
            'net_cashflow': float(
                (payments.aggregate(Sum('amount'))['amount__sum'] or 0) - 
                (loans.aggregate(Sum('amount'))['amount__sum'] or 0)
            ),
        }
        
        # 🔥 Daily table data
        daily_data = []
        for i in range((end_date - start_date).days + 1):
            current_date = start_date + timedelta(days=i)
            day_loans = loans.filter(loan_date=current_date)
            day_payments = payments.filter(payment_date=current_date)
            
            daily_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'total_loans': float(day_loans.aggregate(Sum('amount'))['amount__sum'] or 0),
                'loan_count': day_loans.count(),
                'total_payments': float(day_payments.aggregate(Sum('amount'))['amount__sum'] or 0),
                'payment_count': day_payments.count(),
                'net_cashflow': float(
                    (day_payments.aggregate(Sum('amount'))['amount__sum'] or 0) - 
                    (day_loans.aggregate(Sum('amount'))['amount__sum'] or 0)
                )
            })
        
        return JsonResponse({
            'success': True,
            'summary': summary,
            'table_data': daily_data[:50]
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

def report_page(request):
    return render(request, 'report/report.html')

@csrf_exempt
@require_http_methods(["POST"])
def cashflow_report(request):
    try:
        data = json.loads(request.body)
        print(f"🔥 RAW API DATA: {data}")
        
        # 🔥 FIX: String → Int conversion
        district_id = int(data.get('district')) if data.get('district') else None
        municipality_id = int(data.get('municipality')) if data.get('municipality') else None
        barangay_id = int(data.get('barangay')) if data.get('barangay') else None
        
        print(f"🔍 FILTERS: D={district_id}, M={municipality_id}, B={barangay_id}")
        
        # 🔥 BASE QUERY
        loans_qs = LoanTransaction.objects.all()
        payments_qs = Payment.objects.all()
        
        # 🔥 DISTRICT FILTER
        if district_id is not None:
            loans_qs = loans_qs.filter(district_id=district_id)
            payments_qs = payments_qs.filter(
                Q(district_id=district_id) | 
                Q(beneficiary__barangay__municipality__district_id=district_id)
            )
            print(f"✅ District {district_id}: {loans_qs.count()} loans")
        
        # 🔥 MUNICIPALITY FILTER
        if municipality_id is not None:
            loans_qs = loans_qs.filter(municipality_id=municipality_id)
            payments_qs = payments_qs.filter(
                Q(municipality_id=municipality_id) | 
                Q(beneficiary__barangay__municipality_id=municipality_id)
            )
            print(f"✅ Municipality {municipality_id}: {loans_qs.count()} loans")
        
        # 🔥 BARANGAY FILTER
        if barangay_id is not None:
            loans_qs = loans_qs.filter(barangay_id=barangay_id)
            payments_qs = payments_qs.filter(
                Q(barangay_id=barangay_id) | 
                Q(beneficiary__barangay_id=barangay_id)
            )
            print(f"✅ Barangay {barangay_id}: {loans_qs.count()} loans")
        
        # 🔥 TOTALS
        total_loans = loans_qs.aggregate(Sum('amount'))['amount__sum'] or 0
        total_payments = payments_qs.aggregate(Sum('amount'))['amount__sum'] or 0
        
        print(f"💰 FINAL: Loans=₱{total_loans}, Payments=₱{total_payments}")
        
        return JsonResponse({
            'summary': {
                'total_loans': float(total_loans),
                'total_payments': float(total_payments),
                'net_cashflow': float(total_payments - total_loans)
            }
        })
    except Exception as e:
        print(f"❌ ERROR: {e}")
        return JsonResponse({'error': str(e)}, status=500)
    
def export_cashflow_excel(request):
    # Parse filters safely
    district_id = request.GET.get('district')
    municipality_id = request.GET.get('municipality')
    barangay_id = request.GET.get('barangay')
    
    # Filter data
    loans_qs = LoanTransaction.objects.all()
    payments_qs = Payment.objects.all()
    
    try:
        if district_id: 
            loans_qs = loans_qs.filter(district_id=int(district_id))
            payments_qs = payments_qs.filter(Q(district_id=int(district_id)))
        if municipality_id:
            loans_qs = loans_qs.filter(municipality_id=int(municipality_id))
            payments_qs = payments_qs.filter(Q(municipality_id=int(municipality_id)))
        if barangay_id:
            loans_qs = loans_qs.filter(barangay_id=int(barangay_id))
            payments_qs = payments_qs.filter(Q(barangay_id=int(barangay_id)))
    except ValueError:
        pass  # Invalid filter IDs
    
    total_loans = loans_qs.aggregate(Sum('amount'))['amount__sum'] or 0
    total_payments = payments_qs.aggregate(Sum('amount'))['amount__sum'] or 0
    net_cashflow = total_payments - total_loans
    
    # 🔥 CREATE EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cashflow Report"
    
    # STYLES
    title_font = Font(bold=True, size=16, color="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_blue = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_green = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    
    # 🔥 TITLE (Safe merge)
    ws['A1'] = f"🏦 CASHFLOW REPORT"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:G1')  # Merge AFTER setting value
    
    # 🔥 FILTER INFO
    filter_text = f"District: {district_id or 'All'} | Generated: {date.today()}"
    ws['A2'] = filter_text
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # 🔥 SUMMARY TABLE
    ws['A4'] = "📊 SUMMARY"
    ws['A4'].font = header_font
    ws['A4'].fill = header_blue
    ws['A4'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A4:B4')
    
    ws['A5'] = "Total Loans"
    ws['B5'] = f"₱{float(total_loans):,.2f}"
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = "Total Payments"
    ws['B6'] = f"₱{float(total_payments):,.2f}"
    ws['A6'].font = Font(bold=True)
    
    ws['A7'] = "NET CASHFLOW"
    ws['B7'] = f"₱{float(net_cashflow):,.2f}"
    ws['A7'].font = Font(bold=True)
    ws['B7'].font = Font(bold=True, size=14, color="10B981" if net_cashflow >= 0 else "EF4444")
    
    # 🔥 LOANS TABLE
    ws['A9'] = "📋 LOANS DETAILS"
    ws['A9'].font = header_font
    ws['A9'].fill = header_green
    ws['A9'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A9:F9')
    
    # Headers
    headers = ['ID', 'Beneficiary', 'Amount', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    # 🔥 DATA ROWS
    for row_idx, loan in enumerate(loans_qs, 11):
        ws.cell(row=row_idx, column=1, value=loan.id)
        ws.cell(row=row_idx, column=2, value=getattr(loan.beneficiary, 'beneficiary_id', 'N/A'))
        ws.cell(row=row_idx, column=3, value=f"₱{float(loan.amount):,.2f}")
        ws.cell(row=row_idx, column=4, value=loan.loan_date.strftime('%Y-%m-%d'))
        
    
    # 🔥 FIXED COLUMN RESIZE (Skip merged cells)
    for col_num in range(1, ws.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_num)
        max_length = 0
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            cell = row[0]
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    # 🔥 DOWNLOAD
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"cashflow_report_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response